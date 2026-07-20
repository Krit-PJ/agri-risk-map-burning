import json, re, math, zipfile, os
from pathlib import Path
from datetime import datetime, date
from collections import Counter, defaultdict
from openpyxl import load_workbook

ROOT = Path('/mnt/data/hs_2025all_primary_update')
XLSX = Path('/mnt/data/2026_KPI_HS_Anlys.xlsx')
OUT = ROOT / 'data' / 'hotspot'
DISTRICT_ORDER = ['เมืองกำแพงเพชร','ไทรงาม','คลองลาน','ขาณุวรลักษบุรี','คลองขลุง','พรานกระต่าย','ลานกระบือ','ทรายทองวัฒนา','ปางศิลาทอง','บึงสามัคคี','โกสัมพีนคร']
ALLOWED_PLANT = {'นาข้าว','อ้อย','ข้าวโพดและไร่หมุนเวียน','เกษตรอื่น ๆ','พื้นที่ป่า','อื่น ๆ'}

def clean(v):
    if v is None: return ''
    return str(v).strip()

def norm_plant(v):
    s = clean(v)
    if not s: return 'อื่น ๆ'
    if s in ('ข้าว','นาข้าว'): return 'นาข้าว'
    if 'ข้าวโพด' in s or 'ไร่หมุนเวียน' in s: return 'ข้าวโพดและไร่หมุนเวียน'
    if 'อ้อย' in s: return 'อ้อย'
    if 'ป่า' in s: return 'พื้นที่ป่า'
    if s.replace(' ','') == 'เกษตรอื่นๆ': return 'เกษตรอื่น ๆ'
    if s.replace(' ','') in ('อื่นๆ','อื่น'): return 'อื่น ๆ'
    return s

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = clean(v)
    if not s: return None
    for fmt in ['%Y-%m-%d','%d/%m/%Y','%Y/%m/%d']:
        try:
            d = datetime.strptime(s, fmt).date()
            return d
        except Exception:
            pass
    return None

def time_text(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%H:%M')
    s = str(v).strip()
    if not s: return ''
    if isinstance(v, (int,float)) and not isinstance(v,bool):
        n = int(v)
        return f'{n//100:02d}:{n%100:02d} น.'
    if re.fullmatch(r'\d{3,4}', s):
        n=int(s); return f'{n//100:02d}:{n%100:02d} น.'
    return s

def parse_latlon(row, idx):
    for key in ['Lat,Ln','lat,lng','lat,lng','Lat,Lon','lat,lon']:
        if key in idx:
            s=clean(row[idx[key]])
            m=re.search(r'(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)', s)
            if m:
                lat=float(m.group(1)); lon=float(m.group(2));
                if -90<=lat<=90 and -180<=lon<=180: return lat,lon
    if 'Maps' in idx:
        s=clean(row[idx['Maps']])
        m=re.search(r'q=(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)', s)
        if m: return float(m.group(1)), float(m.group(2))
    return None, None

def ym_be(d):
    return f'{d.year+543}-{d.month:02d}'

def read_sheet(sheet_name, dataset_year_be, apply_land_agri=True):
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(h) for h in next(rows)]
    idx = {h:i for i,h in enumerate(headers)}
    # accept either Date or YYYY-MM-DD
    date_col = 'YYYY-MM-DD' if 'YYYY-MM-DD' in idx else 'Date'
    raw_rows = 0; kpt_rows = 0; agri_rows = 0; final = []
    dup = 0; skipped_coord = 0; skipped_date = 0
    seen = set()
    for row in rows:
        raw_rows += 1
        province = clean(row[idx.get('Province')]) if 'Province' in idx else ''
        if province != 'กำแพงเพชร':
            continue
        kpt_rows += 1
        land = clean(row[idx.get('LandType')]) if 'LandType' in idx else ''
        if apply_land_agri and land != 'พื้นที่เกษตร':
            continue
        agri_rows += 1
        plant = norm_plant(row[idx.get('PlantType')] if 'PlantType' in idx else '')
        if plant not in ALLOWED_PLANT:
            plant = 'อื่น ๆ'
        d = to_date(row[idx[date_col]])
        if not d:
            skipped_date += 1; continue
        lat, lon = parse_latlon(row, idx)
        if lat is None or lon is None:
            skipped_coord += 1; continue
        hs = clean(row[0]) or f'{sheet_name}_{len(final)+1:06d}'
        if hs in seen:
            dup += 1; continue
        seen.add(hs)
        amphoe = clean(row[idx.get('Amphoe')]) if 'Amphoe' in idx else ''
        tambon = clean(row[idx.get('Tambon')]) if 'Tambon' in idx else ''
        baan = clean(row[idx.get('BaanN')]) if 'BaanN' in idx else ''
        ttxt = time_text(row[idx.get('Time')]) if 'Time' in idx else ''
        maps = clean(row[idx.get('Maps')]) if 'Maps' in idx else f'http://maps.google.com/maps?q={lat},{lon}'
        feature = {
            'type':'Feature',
            'geometry': {'type':'Point','coordinates':[round(lon,6), round(lat,6)]},
            'properties': {
                'hs_id': hs, 'hsID': hs,
                '__date': d.isoformat(), 'acq_date': d.isoformat(), 'Date': d.isoformat(),
                'acq_time': ttxt, 'Time': ttxt,
                'year_be': d.year + 543,
                'season_be': dataset_year_be,
                'dataset_year_be': dataset_year_be,
                'province': province, 'Province': province, '__province': province,
                'district': amphoe, 'Amphoe': amphoe, '__district': amphoe,
                'subdistrict': tambon, 'Tambon': tambon, '__subdistrict': tambon,
                'village': baan, 'BaanN': baan, '__village': baan,
                'land_type': land, 'LandType': land,
                'plant_type': plant, 'PlantType': plant, 'crop_type': plant, 'crop_type_raw': plant, '__crop': plant, '__plant_type': plant,
                'confidence': row[idx.get('Q')] if 'Q' in idx and row[idx.get('Q')] is not None else '',
                'Q': str(row[idx.get('Q')]) if 'Q' in idx and row[idx.get('Q')] is not None else '',
                'source': '2026_KPI_HS_Anlys.xlsx',
                'source_file': '2026_KPI_HS_Anlys.xlsx',
                'source_sheet': sheet_name,
                'filter_rule': 'Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType normalized to canonical groups; de-duplicate by hsID',
                'month': d.month, '__month': d.month, '__day': d.day,
                'month_key': ym_be(d), 'day_key': d.isoformat(),
                'TambonN': clean(row[idx.get('TambonN')]) if 'TambonN' in idx else tambon,
                'AmphoeN': clean(row[idx.get('AmphoeN')]) if 'AmphoeN' in idx else amphoe,
                'ProvinceN': clean(row[idx.get('ProvinceN')]) if 'ProvinceN' in idx else province,
                'Maps': maps,
            }
        }
        final.append(feature)
    qa = qa_from_features(final, dataset_year_be, '2026_KPI_HS_Anlys.xlsx', sheet_name)
    qa.update({
        'raw_rows': raw_rows,
        'province_kamphaeng_phet_rows': kpt_rows,
        'province_and_agri_rows': agri_rows,
        'after_plant_normalization_rows': len(final),
        'duplicates_removed_by_hsID': dup,
        'missing_coordinate_rows_skipped': skipped_coord,
        'missing_date_rows_skipped': skipped_date,
        'filter_columns': ['YYYY-MM-DD' if sheet_name=='2025All' else date_col, 'Time', 'Province', 'Amphoe', 'Tambon', 'LandType', 'PlantType'],
    })
    return {'type':'FeatureCollection','name':f'hotspot_{dataset_year_be}','features':final}, qa

def qa_from_features(features, dataset_year_be, source_file, sheet):
    by_month=Counter(); by_day=Counter(); by_dist=Counter(); by_plant=Counter(); by_land=Counter(); dates=[]
    for f in features:
        p=f['properties'];
        by_month[p['month_key']]+=1; by_day[p['day_key']]+=1; by_dist[p.get('Amphoe','')]+=1; by_plant[p.get('PlantType','')]+=1; by_land[p.get('LandType','')]+=1
        try: dates.append(datetime.strptime(p['day_key'],'%Y-%m-%d').date())
        except: pass
    return {
        'dataset_year_be': dataset_year_be,
        'source_file': source_file,
        'sheet': sheet,
        'final_feature_count': len(features),
        'date_min': min(dates).isoformat() if dates else None,
        'date_max': max(dates).isoformat() if dates else None,
        'by_month': dict(sorted(by_month.items())),
        'by_day_count': len(by_day),
        'by_day': dict(sorted(by_day.items())),
        'by_district': {d: by_dist.get(d,0) for d in DISTRICT_ORDER},
        'by_plant_type': dict(sorted(by_plant.items())),
        'by_land_type': dict(sorted(by_land.items())),
    }

def load_geojson(path):
    with open(path, encoding='utf-8') as f: return json.load(f)

def recompute_qa_for_file(year):
    data=load_geojson(OUT/f'hotspot_{year}.geojson')
    return qa_from_features(data['features'], year, 'existing standardized dataset + 2026_KPI_HS_Anlys.xlsx', f'hotspot_{year}.geojson')

def aggregate_summaries(years):
    records_day={}; records_month={}; records_year={}
    def add_counter(d,k,v): d[k]=d.get(k,0)+v
    for year in years:
        data=load_geojson(OUT/f'hotspot_{year}.geojson')
        for f in data['features']:
            p=f['properties']; day=p.get('day_key'); month=p.get('month_key'); y=str(p.get('year_be') or year)
            dist=p.get('Amphoe') or p.get('district') or ''
            tamb=(p.get('Tambon') or p.get('subdistrict') or '')
            land=p.get('LandType') or p.get('land_type') or ''
            plant=p.get('PlantType') or p.get('plant_type') or ''
            for key,bucket in [(day,records_day),(month,records_month),(y,records_year)]:
                if not key: continue
                rec=bucket.setdefault(key, {'period':key,'count':0,'by_district':{},'by_tambon':{},'by_land_type':{},'by_plant_type':{}})
                rec['count']+=1
                add_counter(rec['by_district'], dist, 1)
                add_counter(rec['by_tambon'], f'{dist}/{tamb}', 1)
                add_counter(rec['by_land_type'], land, 1)
                add_counter(rec['by_plant_type'], plant, 1)
    return (
        {'source':'hotspot_2567.geojson, hotspot_2568.geojson, hotspot_2569.geojson','filter_rule':'Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType canonical groups','level':'day','records':[records_day[k] for k in sorted(records_day)]},
        {'source':'hotspot_2567.geojson, hotspot_2568.geojson, hotspot_2569.geojson','filter_rule':'Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType canonical groups','level':'month','records':[records_month[k] for k in sorted(records_month)]},
        {'source':'hotspot_2567.geojson, hotspot_2568.geojson, hotspot_2569.geojson','filter_rule':'Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType canonical groups','level':'year','records':[records_year[k] for k in sorted(records_year)]},
    )

# Primary update: Sheet 2025All -> dataset 2568
for sheet, year in [('2025All',2568), ('2026All',2569)]:
    geo, qa = read_sheet(sheet, year)
    with open(OUT/f'hotspot_{year}.geojson','w',encoding='utf-8') as f:
        json.dump(geo,f,ensure_ascii=False,separators=(',',':'))
    with open(OUT/f'hotspot_{year}_qa.json','w',encoding='utf-8') as f:
        json.dump(qa,f,ensure_ascii=False,indent=2)

# Recompute QA for 2567 from existing standardized previous data.
for year in [2567]:
    qa = recompute_qa_for_file(year)
    qa.update({'note':'นำข้อมูลที่จัดทำไว้ก่อนหน้านี้มาเพิ่มเติม เนื่องจากไฟล์หลักรอบนี้ใช้ Sheet 2025All สำหรับชุดปี 2568 เป็นหลัก'})
    with open(OUT/f'hotspot_{year}_qa.json','w',encoding='utf-8') as f:
        json.dump(qa,f,ensure_ascii=False,indent=2)

# all-years QA
years={}
for year in [2567,2568,2569]:
    qa=json.load(open(OUT/f'hotspot_{year}_qa.json',encoding='utf-8'))
    years[str(year)]={
        'final_feature_count': qa['final_feature_count'],
        'date_min': qa['date_min'], 'date_max': qa['date_max'],
        'by_month': qa['by_month'],
        'by_district': qa['by_district'],
        'by_plant_type': qa['by_plant_type'],
        'source_file': qa.get('source_file'), 'sheet': qa.get('sheet')
    }
allqa={
    'source_file_primary':'2026_KPI_HS_Anlys.xlsx',
    'primary_sheet':'2025All',
    'supplement':'hotspot_2567.geojson from previous standardized data; 2026All sheet used to keep fiscal 2569 coverage consistent',
    'standard_filter':'Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType canonical groups; de-duplicate by hsID; preserve Province-Amphoe-Tambon-BaanN hierarchy',
    'years': years
}
with open(OUT/'hotspot_all_years_qa.json','w',encoding='utf-8') as f: json.dump(allqa,f,ensure_ascii=False,indent=2)

# summaries
for name, obj in zip(['hotspot_day_summary_2567_2569.json','hotspot_month_summary_2567_2569.json','hotspot_year_summary_2567_2569.json'], aggregate_summaries([2567,2568,2569])):
    with open(OUT/name,'w',encoding='utf-8') as f: json.dump(obj,f,ensure_ascii=False,indent=2)

print(json.dumps(allqa,ensure_ascii=False,indent=2))
