"""Build the 2569 hotspot dataset using the hsPlant2026 workbook definition.

The pivot sheet ``hsPlant2026`` is the reconciliation source. Point geometry and
row-level attributes are read from ``hsAll`` because a pivot table has no rows
that can be mapped. Output is written only when both totals agree exactly.
"""

import argparse
import json
import re
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


PROVINCE = "กำแพงเพชร"
LAND_TYPE = "พื้นที่เกษตร"
PLANT_TYPES = ("นาข้าว", "อ้อย", "เกษตรอื่น ๆ", "อื่น ๆ")
DISTRICT_ORDER = (
    "เมืองกำแพงเพชร", "ไทรงาม", "คลองลาน", "ขาณุวรลักษบุรี",
    "คลองขลุง", "พรานกระต่าย", "ลานกระบือ", "ทรายทองวัฒนา",
    "ปางศิลาทอง", "บึงสามัคคี", "โกสัมพีนคร",
)
MAP_COORDINATES = re.compile(r"[?&]q=(-?\d+(?:\.\d+)?),\s*(-?\d+(?:\.\d+)?)", re.I)


def clean(value):
    return "" if value is None else str(value).strip()


def compact(value):
    return re.sub(r"\s+", "", clean(value))


def normalize_plant(value):
    lookup = {
        "นาข้าว": "นาข้าว",
        "อ้อย": "อ้อย",
        "เกษตรอื่นๆ": "เกษตรอื่น ๆ",
        "อื่นๆ": "อื่น ๆ",
        "อื่น": "อื่น ๆ",
    }
    return lookup.get(compact(value), "")


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, pattern).date()
            return parsed.replace(year=parsed.year - 543) if parsed.year > 2400 else parsed
        except ValueError:
            continue
    return None


def format_time(value):
    if isinstance(value, (datetime, time)):
        return value.strftime("%H:%M")
    return clean(value)


def coordinates_from_maps(value):
    match = MAP_COORDINATES.search(clean(value))
    if not match:
        return None
    lat, lon = float(match.group(1)), float(match.group(2))
    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        return None
    return [round(lon, 6), round(lat, 6)]


def read_reference(workbook):
    if "hsPlant2026" not in workbook.sheetnames:
        raise ValueError("ไม่พบ sheet hsPlant2026")
    rows = list(workbook["hsPlant2026"].iter_rows(values_only=True))
    header_index = next(
        (i for i, row in enumerate(rows) if compact(row[0]) == "จังหวัด" and "นาข้าว" in {compact(v) for v in row}),
        None,
    )
    if header_index is None:
        raise ValueError("ไม่พบหัวตารางจังหวัด/PlantType ใน sheet hsPlant2026")
    headers = [compact(value) for value in rows[header_index]]
    province_row = next((row for row in rows[header_index + 1:] if compact(row[0]) == PROVINCE), None)
    if province_row is None:
        raise ValueError("ไม่พบแถวจังหวัดกำแพงเพชรใน sheet hsPlant2026")
    by_plant = {}
    for plant in PLANT_TYPES:
        column = headers.index(compact(plant))
        by_plant[plant] = int(province_row[column] or 0)
    return {"sheet": "hsPlant2026", "by_plant_type": by_plant, "total": sum(by_plant.values())}


def feature_from_row(row, index, row_number, dataset_year, source_name):
    get = lambda key: row[index[key]] if key in index and index[key] < len(row) else None
    acquired = parse_date(get("Date") if "Date" in index else get("YYYY-MM-DD"))
    coordinates = coordinates_from_maps(get("Maps"))
    if acquired is None or coordinates is None:
        return None, "missing_date" if acquired is None else "missing_coordinates"
    hotspot_id = clean(get("hsID")) or f"hsAll-row-{row_number}"
    plant = normalize_plant(get("PlantType"))
    district, subdistrict = clean(get("Amphoe")), clean(get("Tambon"))
    village = clean(get("BaanN"))
    iso_date = acquired.isoformat()
    month_key = f"{acquired.year + 543}-{acquired.month:02d}"
    properties = {
        "hs_id": hotspot_id, "hsID": hotspot_id,
        "__date": iso_date, "acq_date": iso_date, "Date": iso_date,
        "acq_time": format_time(get("Time")), "Time": format_time(get("Time")),
        "year_be": acquired.year + 543, "season_be": dataset_year, "dataset_year_be": dataset_year,
        "province": PROVINCE, "Province": PROVINCE, "__province": PROVINCE,
        "district": district, "Amphoe": district, "__district": district,
        "subdistrict": subdistrict, "Tambon": subdistrict, "__subdistrict": subdistrict,
        "village": village, "BaanN": village, "__village": village,
        "land_type": LAND_TYPE, "LandType": LAND_TYPE,
        "plant_type": plant, "PlantType": plant, "crop_type": plant,
        "crop_type_raw": clean(get("PlantType")), "__crop": plant, "__plant_type": plant,
        "confidence": get("Q") if get("Q") is not None else "", "Q": clean(get("Q")),
        "source": source_name, "source_file": source_name, "source_sheet": "hsAll",
        "filter_rule": "Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType in [นาข้าว, อ้อย, เกษตรอื่น ๆ, อื่น ๆ]",
        "month": acquired.month, "__month": acquired.month, "__day": acquired.day,
        "month_key": month_key, "day_key": iso_date,
        "TambonN": clean(get("TambonN")) or subdistrict,
        "AmphoeN": clean(get("AmphoeN")) or district,
        "ProvinceN": clean(get("ProvinceN")) or PROVINCE,
        "Maps": clean(get("Maps")),
    }
    return {"type": "Feature", "geometry": {"type": "Point", "coordinates": coordinates}, "properties": properties}, None


def summarize(features):
    counters = {key: Counter() for key in ("month", "day", "district", "plant", "land")}
    dates = []
    for feature in features:
        props = feature["properties"]
        counters["month"][props["month_key"]] += 1
        counters["day"][props["day_key"]] += 1
        counters["district"][props["Amphoe"]] += 1
        counters["plant"][props["PlantType"]] += 1
        counters["land"][props["LandType"]] += 1
        dates.append(props["day_key"])
    return {
        "final_feature_count": len(features), "date_min": min(dates) if dates else None,
        "date_max": max(dates) if dates else None,
        "by_month": dict(sorted(counters["month"].items())),
        "by_day_count": len(counters["day"]), "by_day": dict(sorted(counters["day"].items())),
        "by_district": {name: counters["district"].get(name, 0) for name in DISTRICT_ORDER},
        "by_plant_type": {name: counters["plant"].get(name, 0) for name in PLANT_TYPES},
        "by_land_type": dict(sorted(counters["land"].items())),
    }


def read_details(workbook, dataset_year, source_name):
    if "hsAll" not in workbook.sheetnames:
        raise ValueError("ไม่พบ sheet hsAll ซึ่งจำเป็นสำหรับพิกัดรายจุด")
    rows = workbook["hsAll"].iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    index = {header: i for i, header in enumerate(headers)}
    required = {"hsID", "Province", "Amphoe", "Tambon", "LandType", "PlantType", "Maps"}
    if not required.issubset(index) or not ({"Date", "YYYY-MM-DD"} & set(index)):
        raise ValueError("sheet hsAll ขาดคอลัมน์รายละเอียดที่จำเป็น")
    features, seen = [], set()
    qa = Counter()
    rejected_plants = Counter()
    for row_number, row in enumerate(rows, start=2):
        qa["input_rows"] += 1
        if clean(row[index["Province"]]) != PROVINCE:
            qa["wrong_province"] += 1
            continue
        if clean(row[index["LandType"]]) != LAND_TYPE:
            qa["wrong_land_type"] += 1
            continue
        plant = normalize_plant(row[index["PlantType"]])
        if plant not in PLANT_TYPES:
            qa["wrong_plant_type"] += 1
            rejected_plants[clean(row[index["PlantType"]]) or "(ว่าง)"] += 1
            continue
        hotspot_id = clean(row[index["hsID"]])
        if hotspot_id and hotspot_id in seen:
            qa["duplicates"] += 1
            continue
        if hotspot_id:
            seen.add(hotspot_id)
        feature, error = feature_from_row(row, index, row_number, dataset_year, source_name)
        if error:
            qa[error] += 1
            continue
        features.append(feature)
    return features, qa, rejected_plants


def write_json(path, value, *, compact_json=False):
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as stream:
        json.dump(value, stream, ensure_ascii=False, separators=(",", ":") if compact_json else None, indent=None if compact_json else 2)
    temporary.replace(path)


def refresh_summaries(output_dir, years=(2567, 2568, 2569)):
    buckets = {"day": {}, "month": {}, "year": {}}
    for year in years:
        with (output_dir / f"hotspot_{year}.geojson").open(encoding="utf-8") as stream:
            collection = json.load(stream)
        for feature in collection.get("features", []):
            props = feature.get("properties", {})
            keys = {"day": props.get("day_key"), "month": props.get("month_key"), "year": str(props.get("year_be") or year)}
            district = props.get("Amphoe") or props.get("district") or ""
            subdistrict = props.get("Tambon") or props.get("subdistrict") or ""
            plant = props.get("PlantType") or props.get("plant_type") or ""
            land = props.get("LandType") or props.get("land_type") or ""
            for level, period in keys.items():
                if not period:
                    continue
                record = buckets[level].setdefault(period, {"period": period, "count": 0, "by_district": {}, "by_tambon": {}, "by_land_type": {}, "by_plant_type": {}})
                record["count"] += 1
                for field, key in (("by_district", district), ("by_tambon", f"{district}/{subdistrict}"), ("by_land_type", land), ("by_plant_type", plant)):
                    record[field][key] = record[field].get(key, 0) + 1
    names = {"day": "hotspot_day_summary_2567_2569.json", "month": "hotspot_month_summary_2567_2569.json", "year": "hotspot_year_summary_2567_2569.json"}
    for level, filename in names.items():
        write_json(output_dir / filename, {
            "source": ", ".join(f"hotspot_{year}.geojson" for year in years),
            "filter_rule": "Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType strict per dataset QA",
            "level": level, "records": [buckets[level][key] for key in sorted(buckets[level])],
        })


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("data/hotspot"))
    parser.add_argument("--dataset-year", type=int, default=2569)
    args = parser.parse_args()
    source_name = args.input.name
    workbook = load_workbook(args.input, read_only=True, data_only=True)
    reference = read_reference(workbook)
    features, counters, rejected = read_details(workbook, args.dataset_year, source_name)
    summary = summarize(features)
    reference_match = summary["final_feature_count"] == reference["total"] and summary["by_plant_type"] == reference["by_plant_type"]
    if not reference_match:
        raise RuntimeError(f"ผลกรอง hsAll ไม่ตรงกับ hsPlant2026: {summary['by_plant_type']} != {reference['by_plant_type']}")
    metadata = {
        "province": PROVINCE, "scope": LAND_TYPE, "dataset_year_be": args.dataset_year,
        "source_file": source_name, "detail_sheet": "hsAll", "reference_sheet": "hsPlant2026",
        "filter_order": ["Province", "LandType", "PlantType"], "plant_types": list(PLANT_TYPES),
        "period": f"{summary['date_min']}/{summary['date_max']}", "reference_total": reference["total"],
        "reference_match": True,
    }
    collection = {"type": "FeatureCollection", "name": f"hotspot_{args.dataset_year}", "metadata": metadata, "features": features}
    qa = {
        **metadata, **summary, "reference_by_plant_type": reference["by_plant_type"],
        "raw_rows": counters["input_rows"], "duplicates_removed_by_hsID": counters["duplicates"],
        "missing_coordinate_rows_skipped": counters["missing_coordinates"], "missing_date_rows_skipped": counters["missing_date"],
        "excluded_wrong_province": counters["wrong_province"], "excluded_wrong_land_type": counters["wrong_land_type"],
        "excluded_wrong_plant_type": counters["wrong_plant_type"], "excluded_plant_type_values": dict(rejected.most_common()),
    }
    write_json(args.output_dir / f"hotspot_{args.dataset_year}.geojson", collection, compact_json=True)
    write_json(args.output_dir / f"hotspot_{args.dataset_year}_qa.json", qa)
    all_years_path = args.output_dir / "hotspot_all_years_qa.json"
    all_years = json.load(all_years_path.open(encoding="utf-8")) if all_years_path.exists() else {"years": {}}
    all_years.pop("supplement", None)
    all_years.update({
        "source_file_primary": source_name, "primary_sheet": "hsPlant2026", "detail_sheet": "hsAll",
        "historical_data": "hotspot_2567.geojson and hotspot_2568.geojson preserved from the previous standardized release",
        "standard_filter": "Province=กำแพงเพชร; LandType=พื้นที่เกษตร; PlantType in [นาข้าว, อ้อย, เกษตรอื่น ๆ, อื่น ๆ]",
    })
    all_years.setdefault("years", {})[str(args.dataset_year)] = {
        key: qa[key] for key in ("final_feature_count", "date_min", "date_max", "by_month", "by_district", "by_plant_type", "source_file")
    }
    write_json(all_years_path, all_years)
    refresh_summaries(args.output_dir)
    print(json.dumps({"output": str(args.output_dir), "reference": reference, "summary": summary, "qa": dict(counters)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
